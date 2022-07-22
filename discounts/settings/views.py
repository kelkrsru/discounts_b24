from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.decorators.csrf import csrf_exempt

from openpyxl import Workbook

from core.bitrix24.bitrix24 import ActivityB24
from core.models import Portals
from activities.models import Activity
from volumes.models import Volume
from .models import SettingsPortal
from .forms import SettingsPortalForm


@xframe_options_exempt
@csrf_exempt
def index(request):
    """Метод главной страницы настроек."""
    if request.method == 'POST':
        member_id: str = request.POST['member_id']
    elif request.method == 'GET':
        member_id: str = request.GET.get('member_id')
    else:
        return render(request, 'error.html', {
            'error_name': 'QueryError',
            'error_description': 'Неизвестный тип запроса'
        })

    portal: Portals = get_object_or_404(Portals, member_id=member_id)
    portal.check_auth()

    settings_portal = get_object_or_404(SettingsPortal, portal=portal)

    activities: Activity = Activity.objects.all()

    try:
        activities_installed = ActivityB24(portal).get_all_installed()
    except RuntimeError as ex:
        return render(request, 'error.html', {
            'error_name': ex.args[0],
            'error_description': ex.args[1]
        })

    if 'save-settings' in request.POST:
        form: SettingsPortalForm = SettingsPortalForm(
            request.POST or None,
            instance=settings_portal,
        )
        if form.is_valid():
            fields_form = form.save(commit=False)
            fields_form.portal = portal
            fields_form.save()
    else:
        form: SettingsPortalForm = SettingsPortalForm(
            instance=settings_portal,
        )

    context = {
        'activities': activities,
        'activities_installed': activities_installed,
        'member_id': member_id,
        'form': form,
    }
    return render(request, 'settings/index.html', context)


@xframe_options_exempt
@csrf_exempt
def export_volumes_2_excel(request):
    """Метод экспорта всех объемов в файл excel."""

    volumes = Volume.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.'
                     'spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=report.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Volumes'

    columns = [
        'pk',
        'nom_group_id',
        'company_id',
        'volume'
    ]

    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for volume in volumes:
        row_num += 1

        row = [
            volume.pk,
            volume.nomenclature_group_id,
            volume.company_id,
            volume.volume,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response
